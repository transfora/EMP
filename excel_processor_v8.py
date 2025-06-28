#!/usr/bin/env python3

# -*- coding: utf-8 -*-

"""
Модуль обработки Excel файлов v8.1
Специализированная логика для файлов дислокации вагонов

ИСПРАВЛЕНИЯ v8.1:
- Добавлена поддержка обработки колонок с датами
- Гибкое форматирование дат согласно инструкциям OneDrive
- Обратная совместимость с v8.0

ИСПРАВЛЕНИЯ v8.0:
- Исправлено именование результирующих файлов
- Добавлено цветовое форматирование из OneDrive
- Улучшена статистика обработки
"""

import pandas as pd
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
import tempfile
import os
import shutil
from datetime import datetime
import locale
import re
from logger import get_logger

logger = get_logger(__name__)

class ExcelProcessor:
    """Обработчик Excel файлов v8.1"""
    
    def __init__(self, instructions):
        """Инициализация обработчика"""
        self.instructions = instructions
        self.processing_stats = {
            'processed_rows': 0,
            'applied_rules': 0,
            'created_columns': 0,
            'formatted_date_columns': 0  # Новая статистика v8.1
        }
        
        # Словарь для преобразования названий месяцев
        self.month_names = {
            'ru': {
                1: 'янв', 2: 'фев', 3: 'мар', 4: 'апр', 5: 'май', 6: 'июн',
                7: 'июл', 8: 'авг', 9: 'сен', 10: 'окт', 11: 'ноя', 12: 'дек'
            },
            'ru_full': {
                1: 'января', 2: 'февраля', 3: 'марта', 4: 'апреля', 5: 'мая', 6: 'июня',
                7: 'июля', 8: 'августа', 9: 'сентября', 10: 'октября', 11: 'ноября', 12: 'декабря'
            },
            'en': {
                1: 'Jan', 2: 'Feb', 3: 'Mar', 4: 'Apr', 5: 'May', 6: 'Jun',
                7: 'Jul', 8: 'Aug', 9: 'Sep', 10: 'Oct', 11: 'Nov', 12: 'Dec'
            },
            'en_full': {
                1: 'January', 2: 'February', 3: 'March', 4: 'April', 5: 'May', 6: 'June',
                7: 'July', 8: 'August', 9: 'September', 10: 'October', 11: 'November', 12: 'December'
            }
        }
    
    def process_file(self, file_content, original_filename):
        """Обработка Excel файла с правильным именованием и поддержкой дат v8.1"""
        try:
            logger.info(f"Начинаем обработку файла {original_filename}")
            
            # Сохранение входного файла во временный файл
            with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as temp_input:
                temp_input.write(file_content)
                temp_input_path = temp_input.name
            
            # Чтение Excel файла
            df = pd.read_excel(temp_input_path)
            
            # Определение листа для обработки
            workbook = openpyxl.load_workbook(temp_input_path)
            sheet_name = workbook.sheetnames[0]
            logger.info(f"Обрабатываем лист: {sheet_name}")
            logger.info(f"Найдено столбцов в файле: {len(df.columns)}")
            logger.info(f"Заголовки: {list(df.columns)}")
            
            # Подготовка целевых колонок
            target_columns = [col['target_name'] for col in self.instructions['columns']]
            logger.info(f"Целевые столбцы из инструкции: {target_columns}")
            
            # Создание результирующего DataFrame
            result_df = pd.DataFrame()
            
            # Обработка каждой колонки согласно инструкции
            columns_processed = 0
            for col_config in self.instructions['columns']:
                source_name = col_config['source_name']
                target_name = col_config['target_name']
                action = col_config['action']
                value = col_config['value']
                
                # Новые параметры для дат v8.1
                is_date = col_config.get('is_date', False)
                date_format = col_config.get('date_format', 'DD.MM.YYYY')
                date_locale = col_config.get('date_locale', 'ru')
                
                if action == 'create':
                    # Создание новой колонки
                    if target_name == 'проект':
                        result_df[target_name] = ''
                        logger.info(f"✅ Колонка '{target_name}' создана для заполнения по правилам")
                    elif target_name == 'Экспедитор':
                        fixed_value = value if value else "ООО ТРАНСФОРА"
                        result_df[target_name] = fixed_value
                        logger.info(f"✅ Колонка '{target_name}' создана с фиксированным значением '{fixed_value}'")
                    else:
                        result_df[target_name] = value if value else ''
                        logger.info(f"✅ Колонка '{target_name}' создана с значением '{value}'")
                    
                    self.processing_stats['created_columns'] += 1
                    columns_processed += 1
                    
                elif action == 'copy' or action is None:
                    # Копирование существующей колонки
                    source_col = self._find_column_case_insensitive(df, source_name)
                    if source_col:
                        # Копирование данных
                        result_df[target_name] = df[source_col]
                        
                        # Обработка дат v8.1
                        if is_date:
                            result_df[target_name] = self._format_date_column(
                                result_df[target_name], 
                                date_format, 
                                date_locale
                            )
                            self.processing_stats['formatted_date_columns'] += 1
                            logger.info(f"✅ Колонка '{source_col}' скопирована как '{target_name}' с форматированием дат ({date_format})")
                        else:
                            logger.info(f"✅ Колонка '{source_col}' скопирована как '{target_name}'")
                        
                        columns_processed += 1
                    else:
                        logger.warning(f"⚠️ Колонка '{source_name}' не найдена в файле")
                        result_df[target_name] = ''
                        columns_processed += 1
            
            logger.info(f"Выбрано столбцов для обработки: {columns_processed}")
            
            # Применение правил замены
            self._apply_replace_rules(result_df)
            
            # Обновление статистики
            self.processing_stats['processed_rows'] = len(result_df)
            logger.info(f"Обработано строк данных: {len(result_df)}")
            if self.processing_stats['formatted_date_columns'] > 0:
                logger.info(f"Отформатировано колонок с датами: {self.processing_stats['formatted_date_columns']}")
            
            # Создание выходного файла с правильным именем
            output_path = self._create_output_file_v8(result_df, original_filename)
            
            # Удаление временного файла
            os.unlink(temp_input_path)
            
            logger.info(f"Файл {original_filename} обработан успешно -> {os.path.basename(output_path)}")
            return output_path
            
        except Exception as e:
            logger.error(f"Ошибка при обработке файла {original_filename}: {str(e)}")
            raise
    
    def _format_date_column(self, series, date_format, date_locale='ru'):
        """
        Форматирование колонки с датами v8.1
        
        Args:
            series: pandas Series с датами
            date_format: формат вывода (DD.MM.YYYY, DD MMM YYYY, etc.)
            date_locale: локаль для названий месяцев
        
        Returns:
            pandas Series с отформатированными датами
        """
        try:
            def format_single_date(date_value):
                if pd.isna(date_value) or date_value is None:
                    return ''
                
                try:
                    # Попытка преобразовать в datetime
                    if isinstance(date_value, str):
                        # Попробуем разные форматы парсинга
                        for fmt in ['%Y-%m-%d', '%d.%m.%Y', '%d/%m/%Y', '%m/%d/%Y', '%Y-%m-%d %H:%M:%S']:
                            try:
                                parsed_date = datetime.strptime(date_value, fmt)
                                break
                            except ValueError:
                                continue
                        else:
                            # Если не удалось распарсить, попробуем pandas
                            parsed_date = pd.to_datetime(date_value, errors='coerce')
                            if pd.isna(parsed_date):
                                return str(date_value)  # Возвращаем как есть
                    else:
                        # Уже datetime или Timestamp
                        parsed_date = pd.to_datetime(date_value)
                    
                    if pd.isna(parsed_date):
                        return str(date_value)
                    
                    # Форматирование согласно заданному формату
                    day = parsed_date.day
                    month = parsed_date.month
                    year = parsed_date.year
                    
                    if date_format == 'DD.MM.YYYY':
                        return f"{day:02d}.{month:02d}.{year}"
                    elif date_format == 'DD/MM/YYYY':
                        return f"{day:02d}/{month:02d}/{year}"
                    elif date_format == 'DD-MM-YYYY':
                        return f"{day:02d}-{month:02d}-{year}"
                    elif date_format == 'YYYY-MM-DD':
                        return f"{year}-{month:02d}-{day:02d}"
                    elif date_format == 'MM/DD/YYYY':
                        return f"{month:02d}/{day:02d}/{year}"
                    elif date_format == 'DD MMM YYYY':
                        month_name = self.month_names.get(date_locale, self.month_names['ru'])[month]
                        return f"{day:02d} {month_name} {year}"
                    elif date_format == 'DD MMMM YYYY':
                        month_name = self.month_names.get(f"{date_locale}_full", self.month_names['ru_full'])[month]
                        return f"{day:02d} {month_name} {year}"
                    else:
                        # Формат по умолчанию
                        return f"{day:02d}.{month:02d}.{year}"
                        
                except Exception as e:
                    logger.warning(f"Ошибка форматирования даты '{date_value}': {e}")
                    return str(date_value)  # Возвращаем как есть при ошибке
            
            # Применяем форматирование ко всей серии
            formatted_series = series.apply(format_single_date)
            
            logger.info(f"✅ Отформатировано {len(formatted_series)} дат в формате {date_format}")
            return formatted_series
            
        except Exception as e:
            logger.error(f"Критическая ошибка форматирования дат: {e}")
            return series  # Возвращаем исходную серию при критической ошибке
    
    def _find_column_case_insensitive(self, df, column_name):
        """Поиск колонки без учета регистра"""
        for col in df.columns:
            if str(col).lower().strip() == str(column_name).lower().strip():
                return col
        return None
    
    def _apply_replace_rules(self, df):
        """Применение правил замены с подсчетом статистики"""
        rules_applied = 0
        for rule in self.instructions['replace_rules']:
            column = rule['column']
            find_value = rule['find_value']
            replace_value = rule['replace_value']
            project_value = rule.get('project_value')
            project_value2 = rule.get('project_value2')
            
            # Поиск колонки в DataFrame
            target_col = self._find_column_case_insensitive(df, column)
            if not target_col:
                logger.warning(f"⚠️ Колонка '{column}' для правила замены не найдена")
                continue
            
            # Применение правила замены
            mask = df[target_col].astype(str) == str(find_value)
            affected_rows = mask.sum()
            
            if affected_rows > 0:
                # Замена значений в исходной колонке
                df.loc[mask, target_col] = replace_value
                
                # Установка значения проекта, если указано
                if project_value and 'проект' in df.columns:
                    df.loc[mask, 'проект'] = project_value
                    logger.info(f"✅ Установлено значение проекта '{project_value}' для {affected_rows} строк (значение: {find_value})")
                
                if project_value2 and 'Заявка' in df.columns:
                    df.loc[mask, 'Заявка'] = project_value2
                    logger.info(f"✅ Установлено значение проекта '{project_value2}' для {affected_rows} строк (значение: {find_value})")
                
                rules_applied += 1
                logger.info(f"✅ Правило замены применено: '{find_value}' -> '{replace_value}' в колонке '{target_col}' ({affected_rows} строк)")
            else:
                logger.info(f"ℹ️ Правило замены не применено: значение '{find_value}' не найдено в колонке '{target_col}'")
        
        self.processing_stats['applied_rules'] = rules_applied
        logger.info(f"Применено правил замены: {rules_applied}")
    
    def _create_output_file_v8(self, df, original_filename):
        """
        Создание выходного файла с правильным именованием v8.0
        ИСПРАВЛЕНИЕ: Теперь файл создается с правильным именем, а не временным
        """
        # Генерация имени файла с датой и временем
        now = datetime.now()
        timestamp = now.strftime("%Y-%m-%d-%H-%M")
        output_filename = f"TRANSFORA_dislocation_{timestamp}.xlsx"
        
        # Создание временного файла
        temp_output = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
        temp_output.close()
        
        # Сохранение DataFrame в Excel
        with pd.ExcelWriter(temp_output.name, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='Дислокация', index=False)
            
            # Получение объекта workbook для форматирования
            workbook = writer.book
            worksheet = writer.sheets['Дислокация']
            
            # Применение форматирования v8.0 с цветами
            self._apply_formatting_v8(worksheet, df)
        
        # ИСПРАВЛЕНИЕ v8.0: Создание файла с правильным именем
        temp_dir = os.path.dirname(temp_output.name)
        final_output_path = os.path.join(temp_dir, output_filename)
        
        # Перемещение временного файла в файл с правильным именем
        shutil.move(temp_output.name, final_output_path)
        
        logger.info(f"✅ Файл сохранен с правильным именем: {output_filename}")
        return final_output_path
    
    def _apply_formatting_v8(self, worksheet, df):
        """
        Применение форматирования v8.0 с поддержкой цветов из OneDrive
        """
        # Получение настроек форматирования из инструкций
        formatting = self.instructions.get('formatting', {})
        
        # Настройки по умолчанию
        default_font_name = formatting.get('font_name', 'Calibri')
        default_font_size = int(formatting.get('font_size', 10))
        
        # Цвета заголовков
        header_bg_color = formatting.get('header_background_color', 'DDDDDD')
        header_text_color = formatting.get('header_text_color', '000000')
        
        # Цвета ячеек
        cell_bg_color = formatting.get('cell_background_color', 'FFFFFF')
        
        # Базовый шрифт
        base_font = Font(name=default_font_name, size=default_font_size)
        
        # Применение шрифта ко всем ячейкам
        for row in worksheet.iter_rows():
            for cell in row:
                cell.font = base_font
                cell.alignment = Alignment(vertical='center')
                
                # Цвет фона для обычных ячеек
                if cell_bg_color and cell_bg_color != 'FFFFFF':
                    cell.fill = PatternFill(start_color=cell_bg_color,
                                          end_color=cell_bg_color,
                                          fill_type='solid')
        
        # Автоподбор ширины колонок
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            
            for cell in column:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            
            # Установка ширины с ограничениями
            adjusted_width = min(max(max_length + 2, 10), 50)
            worksheet.column_dimensions[column_letter].width = adjusted_width
        
        # Форматирование заголовков с цветами
        header_font = Font(name=default_font_name, size=default_font_size, bold=True, color=header_text_color)
        header_fill = PatternFill(start_color=header_bg_color, end_color=header_bg_color, fill_type='solid')
        
        for cell in worksheet[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        logger.info(f"✅ Применено форматирование v8.0: шрифт {default_font_name} {default_font_size}pt, цвета из OneDrive")
    
    def get_processing_statistics(self):
        """Получение статистики обработки для email шаблонов v8.1"""
        return self.processing_stats