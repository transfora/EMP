#!/usr/bin/env python3

# -*- coding: utf-8 -*-

"""
Модуль работы с OneDrive v8.1
Загрузка и парсинг файла инструкции

ИСПРАВЛЕНИЯ v8.1:
- Поддержка параметров обработки дат (is_date, date_format, date_locale)
- Расширенная валидация настроек дат
- Обратная совместимость с v8.0

ИСПРАВЛЕНИЯ v8.0:
- Поддержка листа formatting для цветового оформления
- Расширенная обработка email шаблонов
- Валидация структуры и цветов
"""

import requests
import tempfile
import os
from openpyxl import load_workbook
from openpyxl.styles import Color
from logger import get_logger

logger = get_logger(__name__)

class OneDriveHandler:
    """Обработчик файлов OneDrive v8.1"""
    
    def __init__(self, instruction_url):
        """Инициализация обработчика"""
        self.instruction_url = instruction_url
    
    def get_processing_instructions(self):
        """Загрузка и парсинг файла инструкции v8.1"""
        try:
            logger.info("Загружаем инструкцию с OneDrive...")
            
            # Скачивание файла
            response = requests.get(self.instruction_url, timeout=30)
            response.raise_for_status()
            logger.info(f"Файл инструкции загружен, размер: {len(response.content)} байт")
            
            # Сохранение во временный файл
            with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as temp_file:
                temp_file.write(response.content)
                temp_file_path = temp_file.name
            
            logger.info(f"Парсим файл инструкции: {temp_file_path}")
            
            # Парсинг файла v8.1
            instructions = self._parse_instruction_file_v8_1(temp_file_path)
            
            # Удаление временного файла
            os.unlink(temp_file_path)
            
            logger.info("✅ Файл инструкции успешно обработан v8.1")
            return instructions
            
        except Exception as e:
            logger.error(f"Ошибка при обработке файла инструкции: {str(e)}")
            return None
    
    def _parse_instruction_file_v8_1(self, file_path):
        """Парсинг Excel файла инструкции v8.1 с поддержкой дат"""
        try:
            workbook = load_workbook(file_path, read_only=False)  # read_only=False для доступа к стилям
            
            instructions = {
                'columns': [],
                'replace_rules': [],
                'email_template': {},
                'formatting': {},
                'variables': []
            }
            
            # Парсинг листа 'columns' v8.1 с поддержкой дат
            if 'columns' in workbook.sheetnames:
                logger.info("Найден лист 'columns'")
                instructions['columns'] = self._parse_columns_sheet_v8_1(workbook['columns'])
            
            # Парсинг листа 'replace'
            if 'replace' in workbook.sheetnames:
                logger.info("Найден лист 'replace'")
                instructions['replace_rules'] = self._parse_replace_sheet(workbook['replace'])
            
            # Парсинг листа 'email' v8.0 (расширенный)
            if 'email' in workbook.sheetnames:
                logger.info("Найден лист 'email'")
                instructions['email_template'] = self._parse_email_sheet_v8(workbook['email'])
            
            # Парсинг листа 'formatting' v8.0 (новый)
            if 'formatting' in workbook.sheetnames:
                logger.info("Найден лист 'formatting' v8.0")
                instructions['formatting'] = self._parse_formatting_sheet_v8(workbook['formatting'])
            else:
                logger.info("Лист 'formatting' не найден, используются настройки по умолчанию")
                instructions['formatting'] = self._get_default_formatting()
            
            # Парсинг листа 'instructions' (опционально)
            if 'instructions' in workbook.sheetnames:
                logger.info("Найден лист 'instructions'")
                # Можно добавить парсинг документации при необходимости
            
            workbook.close()
            
            # Валидация инструкций v8.1
            self._validate_instructions_v8_1(instructions)
            
            return instructions
            
        except Exception as e:
            logger.error(f"Ошибка при парсинге файла инструкции: {str(e)}")
            return None
    
    def _parse_columns_sheet_v8_1(self, sheet):
        """Парсинг листа 'columns' v8.1 с поддержкой параметров дат"""
        columns = []
        
        # Определяем структуру заголовков
        headers = []
        first_row = list(sheet.iter_rows(min_row=1, max_row=1, values_only=True))[0]
        for i, header in enumerate(first_row):
            if header:
                headers.append((i, str(header).lower().strip()))
        
        logger.info(f"Найдены заголовки в листе 'columns': {[h[1] for h in headers]}")
        
        # Создаем маппинг заголовков
        header_mapping = {}
        for i, header in headers:
            if 'source' in header or header == 'source_name':
                header_mapping['source_name'] = i
            elif 'target' in header or header == 'target_name':
                header_mapping['target_name'] = i
            elif header == 'action':
                header_mapping['action'] = i
            elif header == 'value':
                header_mapping['value'] = i
            elif 'date' in header and 'format' in header:
                header_mapping['date_format'] = i
            elif header == 'is_date' or ('date' in header and ('is' in header or 'flag' in header)):
                header_mapping['is_date'] = i
            elif 'locale' in header or ('date' in header and 'locale' in header):
                header_mapping['date_locale'] = i
        
        logger.info(f"Маппинг заголовков: {header_mapping}")
        
        # Парсинг данных
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if not row or not row[header_mapping.get('source_name', 0)]:
                continue  # Пропускаем пустые строки
            
            column_config = {
                'source_name': row[header_mapping.get('source_name', 0)],
                'target_name': row[header_mapping.get('target_name', 1)] if len(row) > 1 and row[header_mapping.get('target_name', 1)] else row[header_mapping.get('source_name', 0)],
                'action': row[header_mapping.get('action', 2)] if len(row) > 2 and row[header_mapping.get('action', 2)] else 'copy',
                'value': row[header_mapping.get('value', 3)] if len(row) > 3 and row[header_mapping.get('value', 3)] else None
            }
            
            # Новые параметры для дат v8.1
            if 'is_date' in header_mapping:
                is_date_value = row[header_mapping['is_date']] if len(row) > header_mapping['is_date'] else None
                column_config['is_date'] = self._parse_boolean_value(is_date_value)
            else:
                column_config['is_date'] = False
            
            if 'date_format' in header_mapping:
                date_format_value = row[header_mapping['date_format']] if len(row) > header_mapping['date_format'] else None
                column_config['date_format'] = date_format_value if date_format_value else 'DD.MM.YYYY'
            else:
                column_config['date_format'] = 'DD.MM.YYYY'
            
            if 'date_locale' in header_mapping:
                date_locale_value = row[header_mapping['date_locale']] if len(row) > header_mapping['date_locale'] else None
                column_config['date_locale'] = date_locale_value if date_locale_value else 'ru'
            else:
                column_config['date_locale'] = 'ru'
            
            columns.append(column_config)
            
            # Логирование конфигурации колонок с датами
            if column_config['is_date']:
                logger.info(f"✅ Настроена колонка с датами: '{column_config['source_name']}' -> '{column_config['target_name']}' ({column_config['date_format']}, {column_config['date_locale']})")
        
        logger.info(f"Загружено столбцов для обработки: {len(columns)}")
        
        # Подсчет колонок с датами
        date_columns = [col for col in columns if col['is_date']]
        if date_columns:
            logger.info(f"✅ Найдено {len(date_columns)} колонок с датами")
        
        return columns
    
    def _parse_boolean_value(self, value):
        """Парсинг булевых значений из Excel"""
        if value is None:
            return False
        
        if isinstance(value, bool):
            return value
        
        str_value = str(value).lower().strip()
        return str_value in ['true', '1', 'да', 'yes', 'y']
    
    def _parse_replace_sheet(self, sheet):
        """Парсинг листа 'replace'"""
        replace_rules = []
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if row[0] and row[1] and row[2]:  # Все три колонки должны быть заполнены
                rule = {
                    'column': row[0],
                    'find_value': row[1],
                    'replace_value': row[2],
                    'project_value': row[3] if len(row) > 3 and row[3] else None,
                    'project_value2': row[4] if len(row) > 4 and row[4] else None
                }
                replace_rules.append(rule)
        
        logger.info(f"Загружено правил замены: {len(replace_rules)}")
        return replace_rules
    
    def _parse_email_sheet_v8(self, sheet):
        """Парсинг листа 'email' v8.0 с поддержкой расширенных шаблонов"""
        email_template = {}
        variables = []
        
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if row[0] and row[1]:
                param_name = row[0]
                param_value = row[1]
                
                # Обработка специальных параметров v8.0
                if param_name == 'variables':
                    # Парсинг списка доступных переменных
                    variables = [var.strip() for var in param_value.split(',')]
                    email_template['variables'] = variables
                    logger.info(f"Найдены переменные для email: {variables}")
                else:
                    email_template[param_name] = param_value
        
        logger.info(f"Загружено настроек email v8.0: {len(email_template)}")
        
        # Валидация email шаблона
        if 'body_template' in email_template:
            logger.info("✅ Найден полный шаблон body_template v8.0")
        else:
            logger.info("ℹ️ Используется совместимый режим email шаблонов v6.0")
        
        return email_template
    
    def _parse_formatting_sheet_v8(self, sheet):
        """Парсинг листа 'formatting' v8.0 для цветового оформления"""
        formatting = {}
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if row[0] and row[1]:
                param_name = row[0]
                param_value = row[1]
                
                # Валидация цветовых значений
                if 'color' in param_name.lower():
                    param_value = self._validate_color_value(param_value)
                
                formatting[param_name] = param_value
        
        logger.info(f"Загружено настроек форматирования: {len(formatting)}")
        
        # Логирование найденных цветов
        color_settings = {k: v for k, v in formatting.items() if 'color' in k.lower()}
        if color_settings:
            logger.info(f"✅ Цветовые настройки: {color_settings}")
        
        return formatting
    
    def _validate_color_value(self, color_value):
        """Валидация и нормализация цветовых значений"""
        if not color_value:
            return None
        
        # Удаление символа # если есть
        color_str = str(color_value).replace('#', '').upper()
        
        # Проверка длины (должно быть 6 символов для RGB)
        if len(color_str) == 6:
            try:
                # Проверка что это валидный hex
                int(color_str, 16)
                return color_str
            except ValueError:
                logger.warning(f"Неверный формат цвета: {color_value}, используется значение по умолчанию")
                return None
        else:
            logger.warning(f"Неверная длина цвета: {color_value}, используется значение по умолчанию")
            return None
    
    def _get_default_formatting(self):
        """Настройки форматирования по умолчанию"""
        return {
            'font_name': 'Calibri',
            'font_size': '10',
            'header_background_color': 'DDDDDD',
            'header_text_color': '000000',
            'cell_background_color': 'FFFFFF'
        }
    
    def _validate_instructions_v8_1(self, instructions):
        """Валидация структуры инструкций v8.1 с проверкой настроек дат"""
        required_sections = ['columns', 'replace_rules', 'email_template']
        
        for section in required_sections:
            if section not in instructions:
                logger.warning(f"⚠️ Отсутствует обязательный раздел: {section}")
            elif not instructions[section]:
                logger.warning(f"⚠️ Раздел '{section}' пуст")
        
        # Валидация колонок
        if instructions['columns']:
            required_columns = [col for col in instructions['columns'] if col['action'] == 'create']
            logger.info(f"Найдено создаваемых колонок: {len(required_columns)}")
            
            # Валидация настроек дат v8.1
            date_columns = [col for col in instructions['columns'] if col['is_date']]
            if date_columns:
                logger.info(f"✅ Найдено колонок с датами: {len(date_columns)}")
                
                # Проверка корректности форматов дат
                valid_formats = ['DD.MM.YYYY', 'DD/MM/YYYY', 'DD-MM-YYYY', 'YYYY-MM-DD', 'MM/DD/YYYY', 'DD MMM YYYY', 'DD MMMM YYYY']
                for col in date_columns:
                    if col['date_format'] not in valid_formats:
                        logger.warning(f"⚠️ Неизвестный формат даты '{col['date_format']}' для колонки '{col['target_name']}'. Поддерживаются: {valid_formats}")
                    
                    if col['date_locale'] not in ['ru', 'en']:
                        logger.warning(f"⚠️ Неизвестная локаль '{col['date_locale']}' для колонки '{col['target_name']}'. Поддерживаются: ru, en")
        
        # Валидация правил замены
        if instructions['replace_rules']:
            projects = set(rule.get('project_value') for rule in instructions['replace_rules'] if rule.get('project_value'))
            logger.info(f"Найдено уникальных проектов в правилах: {len(projects)} - {projects}")
            
            if instructions['replace_rules']:
                projects2 = set(rule.get('project_value2') for rule in instructions['replace_rules'] if rule.get('project_value2'))
                logger.info(f"Найдено уникальных заявок в правилах: {len(projects2)} - {projects2}")
        
        # Валидация email шаблона
        if instructions['email_template']:
            if 'subject' not in instructions['email_template']:
                logger.warning("⚠️ Не найден параметр 'subject' в email настройках")
            
            # Проверка наличия переменных в шаблоне
            if 'body_template' in instructions['email_template']:
                template = instructions['email_template']['body_template']
                variables_in_template = []
                for var in ['{source_filename}', '{output_filename}', '{processing_date}', '{processed_rows}']:
                    if var in template:
                        variables_in_template.append(var)
                logger.info(f"Переменные в шаблоне: {variables_in_template}")
        
        logger.info("✅ Валидация инструкций v8.1 завершена")